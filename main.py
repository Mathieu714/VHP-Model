import VHP.py as pr
import openpyxl
import random
import matplotlib.pyplot as plt
import numpy as np
from sklearn.metrics import root_mean_squared_error, r2_score, mean_squared_error
import time

st = time.process_time()


# generates artificial data to test model

data1 = [[1, 0.3, 7, 24],
         [5, 0.9, 10, 69],
         [3, 0.1, 19, 32],
         [2, 0.3, 18, 12],
         [1, 0.4, 4, 48],
         [5, 0.9, 23, 92]]

data2 = [[2, 7, 0.1, 46.2],
         [4, 19, 0.5, 123],
         [5, 17, 0.6, 132.2],
         [1, 25, 0.8, 126.6],
         [3, 11, 0.2, 73.4],
         [1, 3, 0.2, 29.4]]

train = []

for i in range(0, 20):
    a = random.randint(1, 5)
    b = random.randint(1, 100)
    c = random.random()
    y = 10*(a**2) + 3*(b**2) + 52*(c**2)
    new_data = [a, b, c, y]
    train.append(new_data)

for i in range(0, 20):
    a = random.randint(1, 5)
    b = random.randint(1, 100)
    c = random.random()
    y = 10*a + 3*b + 52*c
    new_data = [a, b, c, y]
    data2.append(new_data)

test2 = [[5, 3, 0.7, 95.4]]
test_train = [[5, 3, 0.7, 302.48]]

var1 = []
var2 = []
var3 = []
y = []

for items in train:
    var1.append(items[0])
    var2.append(items[1])
    var3.append(items[2])
    y.append(items[3])



# import formal clinical data from dataset
path = r'E:\Data\clinical.xlsx'
file = r'E:\Data\Data final'

wb = openpyxl.load_workbook(path)
sheet = wb.active

max_row = sheet.max_row
excel = []
data = []

for i in range(2, max_row+1):
    c2 = sheet.cell(row=i, column=2)
    age = float(c2.value)

    c6 = sheet.cell(row=i, column=6)
    stage = int(c6.value)

    c7 = sheet.cell(row=i, column=7)
    hist = int(c7.value)

    c8 = sheet.cell(row=i, column=8)
    gender = int(c8.value)

    c9 = sheet.cell(row=i, column=9)
    sur_time = int(c9.value)

    c10 = sheet.cell(row=i, column=10)
    death_status = int(c10.value)

    if death_status == 1:
        row_data = [age, stage, hist, gender, sur_time/100]
        data.append(row_data)

print(data)

print(len(data))
train_clinical, test_clinical = pr.train_test_split(data, 0.8)

data, test = train_clinical, test_clinical

# initialize model
lr = 0.5
model = pr.Model(lr=lr, dim=4)


# train model
model.train(data, test, 100)



# visualizing results
total = 0
total_pred = []
total_true = []
for i in range(0, len(test)):
    print(test[i][:-1])
    p = model.predict(test[i][:-1])
    print(p)
    acc = pr.percent_acc(p, test[i][-1])
    print(f'final accuracy is {acc}')
    if acc > 0:
        total += acc
        total_pred.append(p)
        total_true.append(test[i][-1])
    # else:
    #     pass

et = time.process_time()
print(et-st)

best = 0
best_pred = []
best_true = []
for i in range(0, len(test)):
    print(test[i][:-1])
    p = model.predict(test[i][:-1], best=True)
    print(p)
    acc = pr.percent_acc(p, test[i][-1])
    print(f'best accuracy is {acc}')
    if acc > 0:
        best += acc
        best_pred.append(p)
        best_true.append(test[i][-1])
    # else:
    #     pass

print('---------model evaluation----------')

print(f'final total acc is {total/len(test)}')
print(f'best total acc is {best/len(test)}')
weight_array, power_array = model.augmented_array
print('final statement')
for k in range(0, len(weight_array)):
    print(f'{weight_array[k]} x^ {power_array[k]}')

weight_array_best, power_array_best = model.best_weight
print('best statement')
for k in range(0, len(weight_array_best)):
    print(f'{weight_array_best[k]} x^ {power_array_best[k]}')

total_rmse = r2_score(total_true, total_pred)
best_rmse = r2_score(best_true, best_pred)
print(f'total rmse is {total_rmse}')
print(f'best rmse is {best_rmse}')


history = model.history
best_history = model.best_history



x = np.linspace(0, len(history), len(history))
x_best = np.linspace(0, len(best_history), len(best_history))
plt.subplot(121)
plt.scatter(x, history)
plt.xlabel('empower')
plt.ylabel('immediate validation acc')
plt.ylim(0, 1)

plt.subplot(122)
plt.scatter(x_best, best_history)
plt.xlabel('empower')
plt.ylabel('best validation acc')
plt.ylim(0, 1)

plt.show()



